import os, io, time, shutil, asyncio, fitz
from concurrent.futures import ThreadPoolExecutor
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image
from PyPDF2 import PdfReader, PdfWriter
import yadisk

# Конфигурация
YANDEX_TOKEN = "y0__xDG74akqveAAhjblgMggf6o1xKS9mEBdAxviZ1aAoqtrPku2rA5qA"
YANDEX_ID = "1130000065607622"
REMOTE_PDF_PATH = "/input/34329.pdf"
LOCAL_PDF_PATH = "../pdfs/34329.pdf"
PDF_OUTPUT_FOLDER = "../pdfs"
EXCEL_OUTPUT_PATH = "out.xlsx"
CROPS = [(25, 90, 400, 200), (410, 90, 600, 200)]
COLUMN_WIDTHS = [20, 54, 20, 20, 20, 28, 10, 10, 300]

def save_to_yandex_disk(y, folder, file_path):
    filename = os.path.basename(file_path)
    name, ext = os.path.splitext(filename)
    if not y.exists(folder):
        y.mkdir(folder)
    i = 1
    while y.exists(f"{folder}/{filename}"):
        filename = f"{name} _{i}{ext}"
        i += 1
    y.upload(file_path, f"{folder}/{filename}")
    print(f"✅ Загружено: {filename}")

async def async_save_to_yandex_disk(y, folder, file_path, loop, executor):
    await loop.run_in_executor(executor, save_to_yandex_disk, y, folder, file_path)

def crop_image_by_pixels(pdf_path, crops):
    doc = fitz.open(pdf_path)
    image = Image.open(io.BytesIO(doc[0].get_pixmap(dpi=75).tobytes("png")))
    parts = []
    for l, t, r, b in crops:
        buf = io.BytesIO()
        image.crop((l, t, r, b)).save(buf, format='PNG')
        parts.append(buf)
    return parts

def split_pdf_by_pages(input_pdf, out_folder, chunk_size=2):
    base = os.path.splitext(os.path.basename(input_pdf))[0]
    reader = PdfReader(input_pdf)
    os.makedirs(out_folder, exist_ok=True)
    for i in range(0, len(reader.pages), chunk_size):
        writer = PdfWriter()
        for page in reader.pages[i:i+chunk_size]:
            writer.add_page(page)
        with open(os.path.join(out_folder, f"{base}_{i//chunk_size}.pdf"), "wb") as f:
            writer.write(f)
    os.remove(input_pdf)

def generate_yadisk_link(filename):
    base_url = "https://docs.yandex.ru/docs/view?url=ya-disk%3A%2F%2F%2Fdisk%2Foutput%2F"
    return f"{base_url}{filename}&name={filename}&uid={YANDEX_ID}"


def prepare_excel(ws, col_widths):
    headers = ["Id работы", "Автор", "Фамилия", "Имя", "Кодовое слово",
    "Результат", "Орф.\n ошибки", "Пункт.\n ошибки", "Скан работы"]
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = col_widths[col-1]
    ws.row_dimensions[1].height = 50

async def insert_images_to_excel(y, pdf_folder, excel, crops, col_widths, loop, executor):
    wb = Workbook()
    ws = wb.active
    prepare_excel(ws, col_widths)
    row = 2
    tasks = []

    for filename in sorted(os.listdir(pdf_folder)):
        if not filename.lower().endswith(".pdf"):
            continue

        pdf_path = os.path.join(pdf_folder, filename)
        tasks.append(async_save_to_yandex_disk(y, "/output", pdf_path, loop, executor))

        cell = ws.cell(row=row, column=1, value=os.path.splitext(filename)[0])
        cell.alignment = Alignment(horizontal="center", vertical="center")

        images = crop_image_by_pixels(pdf_path, crops)
        for col, img_buf in enumerate(images, start=2):
            col = 6 if col == 3 else col
            ws.add_image(XLImage(img_buf), ws.cell(row=row, column=col).coordinate)
        ws.row_dimensions[row].height = 84

        link = generate_yadisk_link(filename)
        cell = ws.cell(row=row, column=9, value=link)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        row += 1
    wb.save(excel)
    tasks.append(async_save_to_yandex_disk(y, "/xlsx", excel, loop, executor))
    return tasks

async def main():
    start = time.time()
    print("⬇️ Старт обработки...")

    shutil.rmtree(PDF_OUTPUT_FOLDER, ignore_errors=True)
    os.makedirs(PDF_OUTPUT_FOLDER, exist_ok=True)

    y = yadisk.YaDisk(token=YANDEX_TOKEN)
    y.download(REMOTE_PDF_PATH, LOCAL_PDF_PATH)

    split_pdf_by_pages(LOCAL_PDF_PATH, PDF_OUTPUT_FOLDER)

    loop = asyncio.get_event_loop()
    executor = ThreadPoolExecutor()
    upload_tasks = await insert_images_to_excel(
        y, PDF_OUTPUT_FOLDER, EXCEL_OUTPUT_PATH,
        CROPS,COLUMN_WIDTHS,loop,executor
    )

    await asyncio.gather(*upload_tasks)

    print(f"⏱ Выполнено за {time.time() - start:.2f} секунд")


if __name__ == "__main__":
    asyncio.run(main())
