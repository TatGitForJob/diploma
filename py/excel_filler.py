import os, io, fitz
from openpyxl.styles import Alignment, Font
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image

# Конфигурация
YANDEX_TOKEN = "y0__xDG74akqveAAhjblgMggf6o1xKS9mEBdAxviZ1aAoqtrPku2rA5qA"
YANDEX_ID = "1130000065607622"
REMOTE_PDF_PATH = "/input/34329.pdf"
LOCAL_PDF_PATH = "../pdfs/34329.pdf"
PDF_OUTPUT_FOLDER = "../pdfs"
EXCEL_OUTPUT_PATH = "out.xlsx"
CROPS = [(25, 90, 400, 200), (410, 90, 600, 200)]
COLUMN_WIDTHS = [20, 54, 20, 20, 20, 28, 10, 10, 30]


def prepare_excel(ws):
    headers = ["Id работы", "Автор", "Фамилия", "Имя", "Кодовое слово",
    "Результат", "Орф.\n ошибки", "Пункт.\n ошибки", "Скан работы"]
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = COLUMN_WIDTHS[col-1]
    ws.row_dimensions[1].height = 50

def natural_sort_key(filename):
    return int(filename.split('_')[1].split('.')[0])

def fill_text_cells(ws, row, filename):
    cell = ws.cell(row=row, column=1, value=os.path.splitext(filename)[0])
    cell.alignment = Alignment(horizontal="center", vertical="center")

    cell = ws.cell(row=row, column=9, value="Ссылка на PDF")
    cell.style = "Hyperlink"
    cell.alignment = Alignment(vertical="center")

def fill_image_cells(ws, row, filepath):
    images = crop_image_by_pixels(filepath)
    for col, img_buf in enumerate(images, start=2):
        col = 6 if col == 3 else col
        ws.add_image(XLImage(img_buf), ws.cell(row=row, column=col).coordinate)
    ws.row_dimensions[row].height = 84

def crop_image_by_pixels(pdf_path):
    doc = fitz.open(pdf_path)
    image = Image.open(io.BytesIO(doc[0].get_pixmap(dpi=75).tobytes("png")))
    parts = []
    for l, t, r, b in CROPS:
        buf = io.BytesIO()
        image.crop((l, t, r, b)).save(buf, format='PNG')
        parts.append(buf)
    return parts
