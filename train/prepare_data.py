import os, fitz, io, shutil, re
import pandas as pd
from openpyxl import load_workbook
from PIL import Image
import cv2
import numpy as np
from tqdm import tqdm
from openpyxl.drawing.image import Image as XLImage

# Конфигурация
LOCAL_XLSX_DIR = "Moscow_xlsx/done"
IMG_SAVE_ROOT = "./part_1"
CSV_OUTPUT = "./part_1/all_files_final.csv"

error_word=0
all_word=0


# Категории и соответствующие crop-функции
CROPS = {
    "surnames": lambda img: img.crop((0,0, img.width, img.height // 3)),
    "names": lambda img: img.crop((0, img.height // 3 , img.width , 2 * img.height // 3)),
    "codewords": lambda img: img.crop((0, 2 * img.height // 3 , img.width - 540, img.height)),
}


def is_valid_text_field(text):
    return (
        isinstance(text, str)
        and text.strip()
        and all(x not in text for x in [' ', ',', '?'])
        and not re.search(r'[A-Za-z]', text)  # латинские буквы
    )

def process_single_file(xlsx_path, save_root, row_id_start=0, word_id_start=0,all_word=0,error_word=0):
    wb = load_workbook(xlsx_path)
    ws = wb.active

    csv_rows = []
    row_id = row_id_start
    word_id = word_id_start

    for i, row in enumerate(ws.iter_rows(min_row=2), start=0):
        text_fields = {
            "surnames": row[2].value,
            "names": row[3].value,
            "codewords": row[4].value
        }
        if not all(is_valid_text_field(text_fields[key]) for key in text_fields):
            print(f"Пропущена строка {row[0].row} ({xlsx_path})")
            continue
        all_word+=3
        pdf_disk_path = row[0].value  # путь на Яндекс.Диске из первой колонки
        if not isinstance(pdf_disk_path, str) or not pdf_disk_path.strip():
            print(f"Нет пути к PDF в строке {row[0].row} ({xlsx_path})")
            continue

        try:
            image = extract_blue_text(f"Moscow_pdf/{pdf_disk_path.split("_")[0]}/{pdf_disk_path}.pdf")
        except Exception as e:
            print(f"Ошибка при обработке PDF {pdf_disk_path}: {e}")
            continue

        for category, text in text_fields.items():
            word_true = text.strip().lower()
            row_crop = CROPS[category](image)

            debug_path = os.path.join("debug_problem_words", f"{word_true}_{word_id}")

            letter_images = split_word_image_into_letters(row_crop, debug_dir=debug_path)


            if len(letter_images) != len(word_true):
                error_word+=1
                print(f"Несоответствие: '{word_true}' → {len(letter_images)} изображений, {len(word_true)} букв ({f"Moscow_pdf/{pdf_disk_path.split("_")[0]}/{pdf_disk_path}.pdf"})")
                continue
            else:
                shutil.rmtree(debug_path, ignore_errors=True)

            letter_dir = os.path.join(save_root, category, f"{word_true}_{word_id}")
            os.makedirs(letter_dir, exist_ok=True)
            for letter_idx, (letter, letter_img) in enumerate(zip(word_true, letter_images)):
                letter_file = f"{letter}_{letter_idx}.jpg"
                full_path = os.path.join(letter_dir, letter_file)
                letter_img.save(full_path)

                # Добавляем в CSV
                csv_rows.append({
                    "id": row_id,
                    "letter_position": letter_idx,
                    "category": category,
                    "letter": letter,
                    "new_path": full_path,
                    "word_id": word_id,
                    "word_true": word_true
                })
                row_id += 1

            word_id += 1

    return csv_rows, row_id, word_id, all_word, error_word

def split_word_image_into_letters(image, debug_dir="debug_cc"):
    import os, shutil
    shutil.rmtree(debug_dir, ignore_errors=True)
    os.makedirs(debug_dir, exist_ok=True)

    image_gray = image.convert("L")
    image_np = np.array(image_gray)

    binary = cv2.adaptiveThreshold(
        image_np, 255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY_INV,
        blockSize=15,
        C=12
    )
    cv2.imwrite(os.path.join(debug_dir, "1_adaptive_thresh.png"), binary)

    contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    letter_images = []
    bounding_boxes = []

    for cnt in contours:
        x, y, w, h = cv2.boundingRect(cnt)
        area = w * h
        if w < 2 or h < 5 or area < 10:
            continue
        if (w < 25 or h < 25) or w > 120 or h > 120:
            continue

        box = (x, y, x + w, y + h)
        bounding_boxes.append(box)

    bounding_boxes = sorted(bounding_boxes, key=lambda b: b[0])

    for idx, (x1, y1, x2, y2) in enumerate(bounding_boxes):
        cropped = image.crop((x1, y1, x2, y2))
        letter_images.append(cropped)
    
        # Склейка всех букв с разделителями
    if letter_images:
        separator_width = 5
        separator_color = (200, 200, 200)

        total_width = sum(img.width for img in letter_images) + separator_width * (len(letter_images) - 1)
        max_height = max(img.height for img in letter_images)
        combined = Image.new("RGB", (total_width, max_height), (255, 255, 255))

        x_offset = 0
        for i, img in enumerate(letter_images):
            combined.paste(img, (x_offset, 0))
            x_offset += img.width

            if i < len(letter_images) - 1:
                # Добавляем вертикальный разделитель
                separator = Image.new("RGB", (separator_width, max_height), separator_color)
                combined.paste(separator, (x_offset, 0))
                x_offset += separator_width

        combined.save(os.path.join(debug_dir, "combined_letters.png"))

    
    return letter_images


def process_all_excels(xlsx_dir, save_root, output_csv):
    all_csv_rows = []
    row_id = 0
    word_id = 0
    count = 0
    error_word=0
    all_word=0

    for file in os.listdir(xlsx_dir):
        if file.endswith(".xlsx"):
            path = os.path.join(xlsx_dir, file)
            print(f"Обработка файла: {file}")
            rows, row_id, word_id,all_word,error_word = process_single_file(path, save_root, row_id, word_id,all_word,error_word)
            all_csv_rows.extend(rows)
            count += 1
            if count % 10 == 0:
                print(f"Обработано файлов: {count}")

    df_csv = pd.DataFrame(all_csv_rows)
    df_csv.to_csv(output_csv, index=False)
    print(f"\nГотово. CSV: {output_csv}, изображения сохранены в: {save_root}")
    print(all_word," ",error_word)

def extract_blue_text(input_path: str) -> Image.Image:
    image,cropeed_white = prepare_cropped_image(input_path)
    if  cropeed_white == 0:
        raise ValueError(f"Невозможно обрезать pdf по границе: {input_path}")

    cropped = image.crop((int(image.width * 0.025), int(image.height * 0.07), int(image.width * 0.67), int(image.height * 0.215)))    
    image_np = np.array(cropped)

    # Преобразуем в цветовое пространство HSV для лучшего выделения синего
    hsv = cv2.cvtColor(image_np, cv2.COLOR_RGB2HSV)

    # Диапазон синего цвета (можно адаптировать)
    lower_blue = np.array([90, 50, 50])
    upper_blue = np.array([130, 255, 255])

    # Маска синего цвета
    mask = cv2.inRange(hsv, lower_blue, upper_blue)

    if cv2.countNonZero(mask) / mask.shape[0] * mask.shape[1] < 0.005:  # меньше 0.1% синих пикселей
        raise ValueError(f"Синий текст не найден или слишком слабый: {input_path}")

    # Немного увеличим области, чтобы лучше захватить переходящие пиксели
    mask = cv2.dilate(mask, np.ones((2, 2), np.uint8), iterations=1)

    # Создаем черный фон и белый текст
    result = np.zeros_like(cropped)
    result[mask > 0] = [255, 255, 255]  # Белый текст

    # Переводим в оттенки серого
    return Image.fromarray(cv2.cvtColor(result, cv2.COLOR_BGR2GRAY))

def prepare_cropped_image(pdf_path: str):
    import fitz
    import io
    from PIL import Image
    import numpy as np

    white_threshold = 250
    min_nonwhite_rows = 5

    def find_first_content_row(image):
        gray = image.convert("L")
        pixels = np.array(gray)
        h, w = pixels.shape

        for y in range(h):
            row = pixels[y]
            if np.mean(row) < white_threshold:
                count = 0
                for i in range(y, min(y + min_nonwhite_rows, h)):
                    if np.mean(pixels[i]) < white_threshold:
                        count += 1
                if count == min_nonwhite_rows:
                    return y
        return 0

    doc = fitz.open(pdf_path)
    img = Image.open(io.BytesIO(doc[0].get_pixmap(dpi=300).tobytes("png")))

    # Обрезка по краям
    img = img.crop((30,30,img.width - 30,img.height - 30))

    # Поиск начала контента
    crop_y = find_first_content_row(img)
    cropped_img = img.crop((0, crop_y, img.width, img.height))

    return cropped_img , crop_y

shutil.rmtree(IMG_SAVE_ROOT, ignore_errors=True)
shutil.rmtree("debug_problem_words", ignore_errors=True)
os.makedirs(IMG_SAVE_ROOT, exist_ok=True)
process_all_excels(LOCAL_XLSX_DIR, IMG_SAVE_ROOT, CSV_OUTPUT)
