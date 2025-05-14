import os, fitz, io, shutil, re
import pandas as pd
from openpyxl import load_workbook
from PIL import Image
import cv2
import numpy as np
from tqdm import tqdm
from openpyxl.drawing.image import Image as XLImage

# Конфигурация
LOCAL_XLSX_DIR = "Moscow_xlsx/done"
IMG_SAVE_ROOT = "./part_1"
CSV_OUTPUT = "./part_1/all_files_final.csv"

error_word=0
all_word=0


# Категории и соответствующие crop-функции
CROPS = {
    "surnames": lambda img: img.crop((0, 10*img.height // 100, img.width, 28*img.height // 100)),
    "names": lambda img: img.crop((0, 39*img.height // 100 , img.width , 59 * img.height // 100)),
    "codewords": lambda img: img.crop((0, 75*img.height // 100, img.width - 540, 92*img.height // 100)),
}


def is_valid_text_field(text):
    return (
        isinstance(text, str)
        and text.strip()
        and all(x not in text for x in [' ', ',', '?'])
        and not re.search(r'[A-Za-z]', text)  # латинские буквы
    )

def process_single_file(xlsx_path, save_root, word_id_start=0, all_word=0, error_word=0):
    wb = load_workbook(xlsx_path)
    ws = wb.active

    csv_rows = []
    word_id = word_id_start

    for i, row in enumerate(ws.iter_rows(min_row=2), start=0):
        text_fields = {
            "surnames": row[2].value,
            "names": row[3].value,
            "codewords": row[4].value
        }

        if not all(is_valid_text_field(text_fields[key]) for key in text_fields):
            print(f"Пропущена строка {row[0].row} ({xlsx_path})")
            continue

        all_word += 3

        pdf_disk_path = row[0].value
        if not isinstance(pdf_disk_path, str) or not pdf_disk_path.strip():
            print(f"Нет пути к PDF в строке {row[0].row} ({xlsx_path})")
            continue

        try:
            image = extract_blue_text(f"Moscow_pdf/{pdf_disk_path.split('_')[0]}/{pdf_disk_path}.pdf")
        except Exception as e:
            print(f"Ошибка при обработке PDF {pdf_disk_path}: {e}")
            continue

        for category, text in text_fields.items():
            word_true = text.strip().lower()
            row_crop = CROPS[category](image)

            first_letter = word_true[0]
            word_dir = os.path.join(save_root, first_letter)
            os.makedirs(word_dir, exist_ok=True)

            word_file = f"{word_true}_{word_id}.jpg"
            word_path = os.path.join(word_dir, word_file)
            row_crop.save(word_path)

            csv_rows.append({
                "id": word_id,
                "new_path": word_path,
                "word_true": word_true
            })
            word_id += 1

    return csv_rows, word_id, all_word, error_word



def process_all_excels(xlsx_dir, save_root, output_csv):
    all_csv_rows = []
    word_id = 0
    count = 0
    error_word=0
    all_word=0

    for file in os.listdir(xlsx_dir):
        if file.endswith(".xlsx"):
            path = os.path.join(xlsx_dir, file)
            print(f"Обработка файла: {file}")
            rows, word_id,all_word,error_word = process_single_file(path, save_root, word_id,all_word,error_word)
            all_csv_rows.extend(rows)
            count += 1
            if count % 10 == 0:
                print(f"Обработано файлов: {count}")

    df_csv = pd.DataFrame(all_csv_rows)
    df_csv.to_csv(output_csv, index=False)
    print(f"\nГотово. CSV: {output_csv}, изображения сохранены в: {save_root}")
    print(all_word," ",error_word)

def extract_blue_text(input_path: str) -> Image.Image:
    image,cropeed_white = prepare_cropped_image(input_path)
    if  cropeed_white == 0:
        raise ValueError(f"Невозможно обрезать pdf по границе: {input_path}")

    cropped = image.crop((int(image.width * 0.025), int(image.height * 0.07), int(image.width * 0.67), int(image.height * 0.215)))    
    image_np = np.array(cropped)

    # Преобразуем в цветовое пространство HSV для лучшего выделения синего
    hsv = cv2.cvtColor(image_np, cv2.COLOR_RGB2HSV)

    # Диапазон синего цвета (можно адаптировать)
    lower_blue = np.array([90, 50, 50])
    upper_blue = np.array([130, 255, 255])

    # Маска синего цвета
    mask = cv2.inRange(hsv, lower_blue, upper_blue)

    if cv2.countNonZero(mask) / mask.shape[0] * mask.shape[1] < 0.005:  # меньше 0.1% синих пикселей
        raise ValueError(f"Синий текст не найден или слишком слабый: {input_path}")

    # Немного увеличим области, чтобы лучше захватить переходящие пиксели
    mask = cv2.dilate(mask, np.ones((2, 2), np.uint8), iterations=1)

    # Создаем черный фон и белый текст
    result = np.zeros_like(cropped)
    result[mask > 0] = [255, 255, 255]  # Белый текст

    # Переводим в оттенки серого
    return Image.fromarray(cv2.cvtColor(result, cv2.COLOR_BGR2GRAY))

def prepare_cropped_image(pdf_path: str):
    import fitz
    import io
    from PIL import Image
    import numpy as np

    white_threshold = 250
    min_nonwhite_rows = 5

    def find_first_content_row(image):
        gray = image.convert("L")
        pixels = np.array(gray)
        h, w = pixels.shape

        for y in range(h):
            row = pixels[y]
            if np.mean(row) < white_threshold:
                count = 0
                for i in range(y, min(y + min_nonwhite_rows, h)):
                    if np.mean(pixels[i]) < white_threshold:
                        count += 1
                if count == min_nonwhite_rows:
                    return y
        return 0

    doc = fitz.open(pdf_path)
    img = Image.open(io.BytesIO(doc[0].get_pixmap(dpi=300).tobytes("png")))

    # Обрезка по краям
    img = img.crop((30,30,img.width - 30,img.height - 30))

    # Поиск начала контента
    crop_y = find_first_content_row(img)
    cropped_img = img.crop((0, crop_y, img.width, img.height))

    return cropped_img , crop_y

shutil.rmtree(IMG_SAVE_ROOT, ignore_errors=True)
os.makedirs(IMG_SAVE_ROOT, exist_ok=True)
process_all_excels(LOCAL_XLSX_DIR, IMG_SAVE_ROOT, CSV_OUTPUT)
